from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from django.template.loader import get_template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from dateutil.relativedelta import relativedelta
import io
from datetime import datetime
from django.db import transaction
from decimal import Decimal, InvalidOperation
from dateutil.relativedelta import relativedelta 
from reportlab.lib.units import inch, cm
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

from .models import Decompte
from datetime import datetime
from dateutil.relativedelta import relativedelta
from decimal import Decimal  # adapte selon ton projet



from .models import (
    MaitreOuvrage, Prestataire, Marche, 
     OrdreService, Decompte, PV, Document,Acompte,Signataire
)
from .forms import (
    MaitreOuvrageForm, PrestataireForm, MarcheForm,  
    OrdreServiceForm, DecompteForm, 
    PVForm, DocumentForm, AcompteForm,SignataireForm
)

@login_required
def dashboard(request):
    """Dashboard with statistics"""
    context = {
        'total_marches': Marche.objects.count(),
        'marches_en_cours': Marche.objects.filter(statut='en_cours').count(),
        'total_prestataires': Prestataire.objects.count(),
        'total_maitre_ouvrage': MaitreOuvrage.objects.count(),
        'montant_total': Marche.objects.aggregate(Sum('montant'))['montant__sum'] or 0,
        'recent_marches': Marche.objects.select_related('prestataire', 'maitre_ouvrage').order_by('-created_at')[:5],
        'recent_decomptes': Decompte.objects.prefetch_related('marche').order_by('-created_at')[:5],
    }
    return render(request, 'markets/dashboard.html', context)

# Maître d'Ouvrage Views
@login_required
def maitre_ouvrage_list(request):
    nom_query = request.GET.get('nom', '')
    responsable_query = request.GET.get('responsable', '')
    maitre_ouvrages = MaitreOuvrage.objects.all()
    
    if nom_query:
        maitre_ouvrages = maitre_ouvrages.filter
        Q(nom__icontains=nom_query) 
           
        

    if responsable_query:
        maitre_ouvrages = maitre_ouvrages.filter
        Q(responsable__icontains=responsable_query)   
        
    
    paginator = Paginator(maitre_ouvrages, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/maitre_ouvrage_list.html', {
        'page_obj': page_obj,
        'nom_query': nom_query,
        'responsable_query': responsable_query
    })

@login_required
def maitre_ouvrage_create(request):
    if request.method == 'POST':
        form = MaitreOuvrageForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Maître d\'ouvrage créé avec succès.')
            return redirect('maitre_ouvrage_list')
    else:
        form = MaitreOuvrageForm()
    return render(request, 'markets/maitre_ouvrage_form.html', {'form': form, 'title': 'Nouveau Maître d\'Ouvrage'})

@login_required
def maitre_ouvrage_update(request, pk):
    maitre_ouvrage = get_object_or_404(MaitreOuvrage, pk=pk)
    if request.method == 'POST':
        form = MaitreOuvrageForm(request.POST, instance=maitre_ouvrage)
        if form.is_valid():
            form.save()
            messages.success(request, 'Maître d\'ouvrage mis à jour avec succès.')
            return redirect('maitre_ouvrage_list')
    else:
        form = MaitreOuvrageForm(instance=maitre_ouvrage)
    return render(request, 'markets/maitre_ouvrage_form.html', {'form': form, 'title': 'Modifier Maître d\'Ouvrage'})

@login_required
def maitre_ouvrage_detail(request, pk):
    maitre = get_object_or_404(MaitreOuvrage, pk=pk)
   
    return render(request, 'markets/maitre_detail.html', {
        'maitre': maitre,
       
    })
@login_required
def maitre_ouvrage_delete(request, pk):
    maitre_ouvrage = get_object_or_404(MaitreOuvrage, pk=pk)
    if request.method == 'POST':
        maitre_ouvrage.delete()
        messages.success(request, 'Maître d\'ouvrage supprimé avec succès.')
        return redirect('maitre_ouvrage_list')
    return render(request, 'markets/confirm_delete.html', {'object': maitre_ouvrage, 'type': 'Maître d\'Ouvrage'})

# Prestataire Views
@login_required
def prestataire_list(request):
    search_query = request.GET.get('search', '')
    prestataires = Prestataire.objects.all()
    
   
    if search_query:
        prestataires = prestataires.filter(nom__icontains=search_query)
           
        
    
    paginator = Paginator(prestataires, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/prestataire_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })

@login_required
def prestataire_create(request):
    if request.method == 'POST':
        form = PrestataireForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Prestataire créé avec succès.')
            return redirect('prestataire_list')
    else:
        form = PrestataireForm()
    return render(request, 'markets/prestataire_form.html', {'form': form, 'title': 'Nouveau Prestataire'})

@login_required
def prestataire_update(request, pk):
    prestataire = get_object_or_404(Prestataire, pk=pk)
    if request.method == 'POST':
        form = PrestataireForm(request.POST, instance=prestataire)
        if form.is_valid():
            form.save()
            messages.success(request, 'Prestataire mis à jour avec succès.')
            return redirect('prestataire_list')
    else:
        form = PrestataireForm(instance=prestataire)
    return render(request, 'markets/prestataire_form.html', {'form': form, 'title': 'Modifier Prestataire'})

@login_required
def prestataire_detail(request, pk):
    prestataire = get_object_or_404(Prestataire, pk=pk)
   
    
    return render(request, 'markets/prestataire_detail.html', {
        'prestataire': prestataire,
        
    })


@login_required
def prestataire_delete(request, pk):
    prestataire = get_object_or_404(Prestataire, pk=pk)
    if request.method == 'POST':
        prestataire.delete()
        messages.success(request, 'Prestataire supprimé avec succès.')
        return redirect('prestataire_list')
    return render(request, 'markets/confirm_delete.html', {'object': prestataire, 'type': 'Prestataire'})

# Marché Views
@login_required
def marche_list(request):
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type', '')
    statut_filter = request.GET.get('statut', '')
    
    marches = Marche.objects.select_related('prestataire', 'maitre_ouvrage').all()
    
    
    if search_query:
        marches = marches.filter(
            Q(numero__icontains=search_query) | 
            Q(objet__icontains=search_query) |
            Q(prestataire__nom__icontains=search_query)
        )
    
    if type_filter:
        marches = marches.filter(type=type_filter)
    
    if statut_filter:
        marches = marches.filter(statut=statut_filter)
    
    paginator = Paginator(marches, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/marche_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'type_filter': type_filter,
        'statut_filter': statut_filter,
        'type_choices': Marche.TYPE_CHOICES,
        'statut_choices': Marche.STATUT_CHOICES,
    })

@login_required
def marche_create(request):
    if request.method == 'POST':
        form = MarcheForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marché créé avec succès.')
            return redirect('marche_list')
    else:
        form = MarcheForm()
    return render(request, 'markets/marche_form.html', {'form': form, 'title': 'Nouveau Marché'})

@login_required
def marche_detail(request, pk):
    marche = get_object_or_404(Marche, pk=pk)
    decomptes = marche.decomptes.all()
    ordres_service = marche.ordres_service.all()
    #decomptes = marche.decomptes.all()
    #pvs = marche.pvs.all()
    #documents = marche.documents.all()
    
    return render(request, 'markets/marche_detail.html', {
        'marche': marche,
        'decomptes': decomptes,
        
        'ordres_service': ordres_service,
        #'decomptes': decomptes,
        #'pvs': pvs,
        #'documents': documents,
    })


@login_required
def marche_update(request, pk):
    marche = get_object_or_404(Marche, pk=pk)
    if request.method == 'POST':
        form = MarcheForm(request.POST, instance=marche)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marché mis à jour avec succès.')
            return redirect('marche_detail', pk=pk)
    else:
        form = MarcheForm(instance=marche)
    return render(request, 'markets/marche_form.html', {'form': form, 'title': 'Modifier Marché'})

@login_required
def marche_delete(request, pk):
    marche = get_object_or_404(Marche, pk=pk)
    if request.method == 'POST':
       marche.delete()
       messages.success(request, 'Marché supprimé avec succès.')
       return redirect('marche_list')
    return render(request, 'markets/confirm_delete.html', {'object': marche, 'type': 'Marché'})

# PDF Generation
@login_required
def generate_marche_pdf(request, pk):
    marche = get_object_or_404(Marche, pk=pk)
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="marche_{marche.numero}.pdf"'
    
    # Create the PDF object
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center alignment
    )
    
    # Add title
    title = Paragraph("CONTRAT DE MARCHÉ PUBLIC", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Market information
    data = [
        ['Numéro du marché:', marche.numero],
        ['Objet:', marche.objet],
        ['Type:', marche.get_type_display()],
        ['Montant:', f"{marche.montant:,.2f} DH"],
        ['Statut:', marche.get_statut_display()],
        ['Date de signature:', marche.date_signature.strftime('%d/%m/%Y') if marche.date_signature else 'Non définie'],
        ['Date de début:', marche.date_debut.strftime('%d/%m/%Y') if marche.date_debut else 'Non définie'],
        ['Date de fin:', marche.date_fin.strftime('%d/%m/%Y') if marche.date_fin else 'Non définie'],
        ['Maître d\'ouvrage:', marche.maitre_ouvrage.responsable],
        ['Prestataire:', marche.prestataire.nom],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Terms and conditions
    terms_title = Paragraph("CONDITIONS GÉNÉRALES", styles['Heading2'])
    elements.append(terms_title)
    elements.append(Spacer(1, 12))
    
    terms = [
        "1. Le présent marché est conclu conformément aux dispositions du decret n°2-22-431 relatif aux marches publics.",
        "2. Les prestations devront être exécutées dans les délais convenus.",
        "3. Le paiement s'effectuera selon les modalités définies dans le marché.",
        "4. Toute modification du marché devra faire l'objet d'un avenant."
    ]
    
    for term in terms:
        p = Paragraph(term, styles['Normal'])
        elements.append(p)
        elements.append(Spacer(1, 6))
    
    elements.append(Spacer(1, 30))
    
    # Signatures
    signature_data = [
        ['Maître d\'ouvrage', 'Prestataire'],
        ['', ''],
        ['Signature:', 'Signature:'],
    ]
    
    signature_table = Table(signature_data, colWidths=[3*inch, 3*inch])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
    ]))
    
    elements.append(signature_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Decompte
from .forms import DecompteForm


def decompte_create(request):
    def calculer_jours_reels_annee(date_debut, date_fin):
        """
        Calcule le nombre de jours réels dans l'année en cours de la période.
        Prend en compte les années bissextiles.
        """
        annee = date_debut.year
        if date_debut.year != date_fin.year:
            # Si la période s'étend sur plusieurs années, prendre l'année de début
            annee = date_debut.year
        
        # Vérifier si c'est une année bissextile
        if (annee % 4 == 0 and annee % 100 != 0) or (annee % 400 == 0):
            return 366
        else:
            return 365
    
    def calculer_jours_par_mois_dans_periode(date_debut, date_fin):
        """
        Calcule le nombre de jours réels pour chaque mois dans la période donnée.
        Retourne le nombre total de jours et un dictionnaire détaillé.
        """
        total_jours = 0
        detail_mois = {}
        
        current_date = date_debut
        
        while current_date <= date_fin:
            annee = current_date.year
            mois = current_date.month
            
            # Début du mois courant
            debut_mois = current_date.replace(day=1)
            
            # Fin du mois courant
            if mois == 12:
                fin_mois = debut_mois.replace(year=annee + 1, month=1, day=1) - relativedelta(days=1)
            else:
                fin_mois = debut_mois.replace(month=mois + 1, day=1) - relativedelta(days=1)
            
            # Calculer les jours dans ce mois pour notre période
            debut_periode_mois = max(current_date, debut_mois)
            fin_periode_mois = min(date_fin, fin_mois)
            
            jours_dans_mois = (fin_periode_mois - debut_periode_mois).days + 1
            
            cle_mois = f"{annee}-{mois:02d}"
            detail_mois[cle_mois] = jours_dans_mois
            total_jours += jours_dans_mois
            
            # Passer au mois suivant
            if mois == 12:
                current_date = current_date.replace(year=annee + 1, month=1, day=1)
            else:
                current_date = current_date.replace(month=mois + 1, day=1)
        
        return total_jours, detail_mois
    
    if request.method == 'POST':
        form = DecompteForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                decompte = form.save(commit=False)
                if decompte.marche and decompte.periode_debut and decompte.periode_fin:
                    # Validation des dates
                    if decompte.periode_debut > decompte.periode_fin:
                        form.add_error('periode_fin', "La date de fin doit être postérieure à la date de début")
                        return render(request, 'markets/decompte_form.html', {'form': form, 'title': 'Créer Décompte'})
                    
                    if not decompte.marche.date_debut:
                        form.add_error(None, "Le marché associé doit avoir une date de début définie")
                        return render(request, 'markets/decompte_form.html', {'form': form, 'title': 'Créer Décompte'})
                    
                    # Montant annuel et application de la TVA
                    montant_annual_ht = decompte.marche.montant_annual or Decimal(0)
                    montant_annual_ttc = (montant_annual_ht * (Decimal('1') + decompte.tva / Decimal('100'))).quantize(Decimal('0.01'))
                    
                    periodicite = decompte.periodicite.lower() if decompte.periodicite else None
                    
                    # Calcul du montant du décompte basé sur la période réelle
                    duree_decompte_jours = (decompte.periode_fin - decompte.periode_debut).days + 1
                    
                    if periodicite and periodicite == 'prorata':
                        # MODE PRORATA : Calcul basé sur le nombre de jours réels avec prise en compte des mois
                        if decompte.marche.date_fin:
                            # Prorata sur la durée totale du marché
                            duree_totale_marche = (decompte.marche.date_fin - decompte.marche.date_debut).days + 1
                            if duree_totale_marche > 0:
                                # Calculer les jours réels dans la période du décompte
                                jours_reels_decompte, detail_mois_decompte = calculer_jours_par_mois_dans_periode(
                                    decompte.periode_debut, decompte.periode_fin
                                )
                                
                                decompte.montant_ttc = (montant_annual_ttc * Decimal(jours_reels_decompte) / Decimal(duree_totale_marche)).quantize(Decimal('0.01'))
                            else:
                                decompte.montant_ttc = Decimal('0')
                        else:
                            # Prorata sur une année avec calcul précis des jours
                            jours_reels_decompte, detail_mois_decompte = calculer_jours_par_mois_dans_periode(
                                decompte.periode_debut, decompte.periode_fin
                            )
                            
                            # Calculer le nombre de jours réels dans l'année de référence
                            jours_reels_annee = calculer_jours_reels_annee(decompte.periode_debut, decompte.periode_fin)
                            
                            decompte.montant_ttc = (montant_annual_ttc * Decimal(jours_reels_decompte) / Decimal(jours_reels_annee)).quantize(Decimal('0.01'))
                    
                    elif periodicite:
                        # MODE PÉRIODICITÉ FIXE : Montant fixe selon la périodicité choisie
                        if periodicite == 'trimestrielle':
                            # Montant pour 1 trimestre complet
                            decompte.montant_ttc = (montant_annual_ttc / Decimal('4')).quantize(Decimal('0.01'))
                        elif periodicite == 'semestrielle':
                            # Montant pour 1 semestre complet
                            decompte.montant_ttc = (montant_annual_ttc / Decimal('2')).quantize(Decimal('0.01'))
                        elif periodicite == 'mensuelle':
                            # Montant pour 1 mois complet
                            decompte.montant_ttc = (montant_annual_ttc / Decimal('12')).quantize(Decimal('0.01'))
                        elif periodicite == 'annuelle':
                            # Montant annuel complet
                            decompte.montant_ttc = montant_annual_ttc
                        else:
                            # Périodicité non reconnue, utiliser le prorata précis par défaut
                            jours_reels_decompte, detail_mois_decompte = calculer_jours_par_mois_dans_periode(
                                decompte.periode_debut, decompte.periode_fin
                            )
                            jours_reels_annee = calculer_jours_reels_annee(decompte.periode_debut, decompte.periode_fin)
                            decompte.montant_ttc = (montant_annual_ttc * Decimal(jours_reels_decompte) / Decimal(jours_reels_annee)).quantize(Decimal('0.01'))
                    
                    else:
                        # Aucune périodicité définie, utiliser le prorata précis par défaut
                        if decompte.marche.date_fin:
                            duree_totale_marche = (decompte.marche.date_fin - decompte.marche.date_debut).days + 1
                            if duree_totale_marche > 0:
                                jours_reels_decompte, detail_mois_decompte = calculer_jours_par_mois_dans_periode(
                                    decompte.periode_debut, decompte.periode_fin
                                )
                                decompte.montant_ttc = (montant_annual_ttc * Decimal(jours_reels_decompte) / Decimal(duree_totale_marche)).quantize(Decimal('0.01'))
                            else:
                                decompte.montant_ttc = Decimal('0')
                        else:
                            jours_reels_decompte, detail_mois_decompte = calculer_jours_par_mois_dans_periode(
                                decompte.periode_debut, decompte.periode_fin
                            )
                            jours_reels_annee = calculer_jours_reels_annee(decompte.periode_debut, decompte.periode_fin)
                            decompte.montant_ttc = (montant_annual_ttc * Decimal(jours_reels_decompte) / Decimal(jours_reels_annee)).quantize(Decimal('0.01'))
                    
                    # Calcul du montant HT à partir du TTC
                    decompte.montant_ht = (decompte.montant_ttc / (Decimal('1') + decompte.tva / Decimal('100'))).quantize(Decimal('0.01'))
                    
                    # Vérification du reste à payer du marché
                    marche = decompte.marche
                    if decompte.montant_ttc > marche.rest_a_payer:
                        form.add_error(None, f"Le montant du décompte ({decompte.montant_ttc} DH) dépasse le reste à payer du marché ({marche.rest_a_payer} DH)")
                        return render(request, 'markets/decompte_form.html', {'form': form, 'title': 'Créer Décompte'})
                    
                    # Mise à jour du reste à payer du marché
                    marche.rest_a_payer -= decompte.montant_ttc
                    marche.save()
                    
                    decompte.save()
                    
                    messages.success(request, f'Décompte créé avec succès. Montant: {decompte.montant_ttc} DH TTC')
                    return redirect('decompte_list')
                    
            except Exception as e:
                messages.error(request, f"Une erreur est survenue : {str(e)}")
                return render(request, 'markets/decompte_form.html', {'form': form, 'title': 'Créer Décompte'})
    else:
        form = DecompteForm()
    
    return render(request, 'markets/decompte_form.html', {
        'form': form, 
        'title': 'Créer Décompte',
        'is_update': False
    })


@login_required
def decompte_create(request):
    form = DecompteForm()

    if request.method == 'POST':
        form = DecompteForm(request.POST)
        
        if form.is_valid():
            try:
                decompte = form.save(commit=False)
               
                
                
                if decompte.marche and decompte.periode_debut and decompte.periode_fin:
                    if decompte.periode_debut > decompte.periode_fin:
                        form.add_error('periode_fin','La date de fin  doit etre postérieure a la date de début')
                        return render (request, 'markets/decompte_form.html',{'form':form,'title': 'Nouveau Décompte'})
                    if not decompte.marche.date_debut:
                        form.add_error(None,"Le marché associé doit avoir unre date de debut definie")
                        return render (request, 'markets/decompte_form.html',{'form': form,'title':"Nouveau Décompte"})
                    

                    montant_annuel = decompte.marche.montant_annual or Decimal(0)
                    periodicite = decompte.marche.periodicite

                    

                    if periodicite == "prorata":
                        jours_periode = (decompte.periode_fin - decompte.periode_debut).days + 1
                        anne = decompte.periode_debut.year
                        jours_annee = 366 if (anne % 4 == 0  and anne % 100 !=0) or (anne % 400==0) else 365
                        montant_calculer = (montant_annuel * Decimal(jours_periode)/Decimal(jours_annee))
                        decompte.montant_ht = montant_calculer.quantize(Decimal('0.01'))
                        print(f"DEBUG - CALCUL Prorata : {jours_periode} jours / {jours_annee} jours = {decompte.montant_ht}")
                    else :
                       periodes_par_an = {
                          'mensuelle' :12,
                          'trimestrielle' :4,
                          'semestrielle' :2
                       }.get(periodicite,1) 

                       decompte.montant_ht = (montant_annuel/Decimal(periodes_par_an)).quantize(Decimal('0.01'))
                       

                       periode_actuelle = decompte.marche.date_debut
                       trouve = False  

                       while periode_actuelle < decompte.periode_fin:
                           periode_suivante = periode_actuelle + relativedelta(months=12//periodes_par_an)

                           if(decompte.periode_debut >= periode_actuelle and decompte.periode_fin <= periode_suivante):
                            trouve = True 
                            break
                           periode_actuelle = periode_suivante
                        
                       if not trouve:
                          
                          form.add_error(None,f"La période ne correspond a aucune periode {periodicite} du marché")
 
                          return render(request, 'markets/decompte_form.html,',{'form' : form, 'title' : 'Nouveau Décompte'})
                
                else:
                    if not decompte.montant_ht:
                      decompte.montant_ht = Decimal("0.00")

                    
                if decompte.montant_ht is not None :
                    decompte.montant_ttc = decompte.montant_ht


                if decompte.statut == 'paye':
                     if not decompte.marche :
                            form.add_error('statut',"Imposible de marquer comme payé sans marché associé ")

                            return render(request, 'markets/decompte_form.html,',{'form': form, 'title':'Nouveau Decompte'})
                        
                     marche = decompte.marche
                     if marche.rest_a_payer < decompte.montant_ttc:
                        form.add_error('montant_ttc', " Le montant dépasse le reste a payer")
                        return render (request, 'markets/decompte_form.html', {'form' : form, 'title': "Nouveau Décompte" })
                     

                     marche.rest_a_payer -= decompte.montant_ttc
                     marche.save()
                decompte.save() 
                messages.success(request, ' Décompte créé avec succés.')
                return redirect ('decompte_list')


                                       

            except Exception as e:
                messages.error(request, f"Une erreur est survenue : {str(e)}")
                return render(request, 'markets/decompte_form.html',{'form': form,'title':"Nouveau Décompte"})
            
        else : 
            return render(request,'markets/decompte_form.html',{'form': form,'title':"Nouveau Décompte"})
    else:
        form = DecompteForm()
       
    return render(request,'markets/decompte_form.html',{'form': form,'title':"Nouveau Décompte"})
        



@login_required
def decompte_update(request, pk):
    decompte = get_object_or_404(Decompte, pk=pk)
    ancien_montant_ttc = decompte.montant_ttc  # Sauvegarder l'ancien montant pour ajuster le reste à payer

    def calculer_jours_reels_annee(date_debut, date_fin):
        annee = date_debut.year
        if (annee % 4 == 0 and annee % 100 != 0) or (annee % 400 == 0):
            return 366
        else:
            return 365

    def calculer_jours_par_mois_dans_periode(date_debut, date_fin):
        total_jours = 0
        detail_mois = {}
        current_date = date_debut

        while current_date <= date_fin:
            annee = current_date.year
            mois = current_date.month
            debut_mois = current_date.replace(day=1)
            fin_mois = (debut_mois.replace(month=mois + 1, day=1) - relativedelta(days=1)) if mois < 12 else (debut_mois.replace(year=annee + 1, month=1, day=1) - relativedelta(days=1))
            debut_periode_mois = max(current_date, debut_mois)
            fin_periode_mois = min(date_fin, fin_mois)
            jours_dans_mois = (fin_periode_mois - debut_periode_mois).days + 1
            cle_mois = f"{annee}-{mois:02d}"
            detail_mois[cle_mois] = jours_dans_mois
            total_jours += jours_dans_mois
            current_date = (current_date.replace(month=mois + 1, day=1)) if mois < 12 else (current_date.replace(year=annee + 1, month=1, day=1))

        return total_jours, detail_mois

    if request.method == 'POST':
        form = DecompteForm(request.POST, request.FILES, instance=decompte)
        if form.is_valid():
            decompte = form.save(commit=False)
            marche = decompte.marche

            # Validation des dates
            if decompte.periode_debut and decompte.periode_fin and decompte.periode_debut > decompte.periode_fin:
                form.add_error('periode_fin', "La date de fin doit être postérieure à la date de début")
                return render(request, 'markets/decompte_form.html', {'form': form, 'title': 'Modifier Décompte'})

            # --- GESTION DU MONTANT TTC EDITABLE ---
            if 'montant_ttc' in form.cleaned_data and form.cleaned_data['montant_ttc'] is not None:
                # L'utilisateur a saisi un montant manuellement
                decompte.montant_ttc = form.cleaned_data['montant_ttc']
            else:
                # Calcul automatique selon prorata / périodicité
                montant_annual_ht = marche.montant_annual or Decimal(0)
                montant_annual_ttc = (montant_annual_ht * (Decimal('1') + decompte.tva / Decimal('100'))).quantize(Decimal('0.01'))

                # Ici tu peux conserver ton ancien calcul prorata / périodicité
                # Exemple pour prorata précis :
                if decompte.periode_debut and decompte.periode_fin:
                    jours_reels_decompte, _ = calculer_jours_par_mois_dans_periode(decompte.periode_debut, decompte.periode_fin)
                    jours_reels_annee = calculer_jours_reels_annee(decompte.periode_debut, decompte.periode_fin)
                    decompte.montant_ttc = (montant_annual_ttc * Decimal(jours_reels_decompte) / Decimal(jours_reels_annee)).quantize(Decimal('0.01'))

            # Calcul du montant HT
            decompte.montant_ht = (decompte.montant_ttc / (Decimal('1') + decompte.tva / Decimal('100'))).quantize(Decimal('0.01'))

            # Ajustement du reste à payer du marché
            difference_montant = decompte.montant_ttc - ancien_montant_ttc
            reste_disponible = marche.rest_a_payer + ancien_montant_ttc
            if decompte.montant_ttc > reste_disponible:
                form.add_error(None, f"Le nouveau montant ({decompte.montant_ttc} DH) dépasse le reste disponible ({reste_disponible} DH)")
                return render(request, 'markets/decompte_form.html', {'form': form, 'title': 'Modifier Décompte'})

            marche.rest_a_payer -= difference_montant
            marche.save()
            decompte.save()

            # Messages utilisateurs
            if difference_montant > 0:
                messages.success(request, f'Décompte modifié avec succès. Montant augmenté de {difference_montant} DH.')
            elif difference_montant < 0:
                messages.success(request, f'Décompte modifié avec succès. Montant diminué de {abs(difference_montant)} DH.')
            else:
                messages.success(request, f'Décompte modifié avec succès. Montant inchangé.')

            return redirect('decompte_list')
    else:
        form = DecompteForm(instance=decompte)

    return render(request, 'markets/decompte_form.html', {
        'form': form,
        'title': 'Modifier Décompte',
        'decompte': decompte,
        'is_update': True
    })
@login_required
def decompte_list(request):
    
    #decompte = Decompte.objects.prefetch_related('marche').all()
    decompte = Decompte.objects.select_related('marche').all()
    #marches = Marche.objects.prefetch_related('decomptes').all()  
    marches = Marche.objects.select_related('decomptes').all()
   
    paginator = Paginator(decompte, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/decompte_list.html', {
        'page_obj': page_obj,          # Paginated Decomptes
        'all_marches': marches,        # All Marches with related Decomptes
        'statut_choices': Decompte.STATUT_CHOICES,
    })

@login_required
def decompte_delete(request, pk):
    decompte = get_object_or_404(Decompte,pk=pk)
    if request.method == 'POST':
        try:
            marche = decompte.marche
            montant_ttc = decompte.montant_ttc
            statut = decompte.statut

            decompte.delete()

            if statut == 'paye' and marche:
                marche.rest_a_payer+=montant_ttc
                marche.save()
                messages.success(request,f'Décompte supprimé avec succes et {montant_ttc} DH réinté grés au reste a payer')
            else:
                messages.success(request,'Décompte supprimé avecsuccés,')
            return redirect('decompte_list')
        except Exception as e :
            messages.error(request,f'Error lors de la suppression  :{str(e)}')
            return redirect('decompte_list')
    return render(request,'markets/confirm_delete.html',{'objet': decompte , 'type' :'Décompte'})



#generate PDF for Decompte
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from openpyxl.utils import get_column_letter
from decimal import Decimal,InvalidOperation
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import get_column_letter


from .models import Decompte  # adapte selon ton projet

@login_required
def generate_decompte_pdf(request,pk):
    import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import Decompte


import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Decompte, Ligne
from openpyxl.drawing.image import Image as XLImage

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER
from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Decompte, Ligne
from io import BytesIO
from num2words import num2words
from decimal import Decimal, ROUND_DOWN
@login_required
def generate_decompte_pdf(request, pk):
    decompte = get_object_or_404(Decompte, pk=pk)
    marche = decompte.marche
    lignes = Ligne.objects.filter(marche=marche)

    # Create the response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="decompte_{decompte.id}.pdf"'

    # PDF setup
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

   

# === LOGOS ===
   
    try:
     logo2 = Image("static/img/LOGO.jpg", width=13*cm, height=3*cm)
    except:
     logo2 = Spacer(3*cm, 3*cm)

    
    header_logos = Table(
    [[logo2]],
    colWidths=[4*cm, 5*cm, 4*cm]  # CENTER logo gets more space
    )

    header_logos.setStyle(TableStyle([
    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
  ]))

    elements.append(header_logos)
    elements.append(Spacer(1, 10))
        
    # Custom styles
    style_center = ParagraphStyle(
        'Center',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=10,
        spaceAfter=6
    )
    
    style_bold = ParagraphStyle(
        'Bold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10
    )

    # === Header ===
    
    #elements.append(Paragraph(header_text, style_center))
    elements.append(Spacer(1, 12))

    # === Market Title ===
    market_title = f"<b>Marché Reconductible N°{marche.numero}  relatif à  : {marche.objet}</b>"
    elements.append(Paragraph(market_title, style_center))
    elements.append(Spacer(1, 12))

    # === Company Information ===
    company_data = [
        ["Société :", f"{marche.prestataire.nom}"],
        ["R.C :", f"{marche.prestataire.numero_registre}"],
        ["CNSS :",f"{marche.prestataire.cnss}"], 
        ["C.B :", f"{marche.prestataire.compte}"],
        ["", ""],
        ["Montant de l'acompte", f"{decompte.montant_ttc:,.2f}"]
    ]
    
    company_table = Table(company_data, colWidths=[4*cm, 11*cm])
    company_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 12))

    # === Decompte Title ===
    decompte_title = f"<b>Décompte {decompte.type} N°{decompte.numero}<br/>Prestations réalisées du {decompte.periode_debut} au {decompte.periode_fin}</b>"
    elements.append(Paragraph(decompte_title, style_center))
    elements.append(Spacer(1, 12))

    # === Main Table ===
    table_data = [["N° des prix", "Désignation", "Unité de mesure", "Qte", "Prix unitaire", "Prix Total"]]

    # Fill table with Lignes
    total_annuel_hors_tva = Decimal(0)
    for ligne in lignes:
        prix_total = ligne.quantite * ligne.prix_unitaire
        total_annuel_hors_tva += prix_total
        table_data.append([
            str(ligne.numero_prix),
            ligne.designation,
            ligne.unite_mesure,
            str(ligne.quantite),
            ligne.prix_unitaire,  # Masked unit price as in the image
            f"{prix_total:.2f}"
        ])

    # Add totals
    tva = total_annuel_hors_tva * Decimal("0.2")
    total_ttc_annuel = total_annuel_hors_tva + tva

    table_data.append(["", "", "", "", "Total Annuel Hors TVA", f"{total_annuel_hors_tva:.2f}"])
    table_data.append(["", "", "", "", "TVA 20 %", f"{tva:.2f}"])
    table_data.append(["", "", "", "", "TOTAL TTC", f"{total_ttc_annuel:.2f}"])
    table_data.append(["", f"TOTAL des Prestations réalisées du {decompte.periode_debut} au {decompte.periode_fin} TTC", "", "", "", f"{decompte.montant_ttc:.2f}"])

    # Style the main table
    table = Table(table_data, colWidths=[2*cm, 7*cm, 2.5*cm, 2*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-4), 0.5, colors.black),  # Grid until annual totals
        ('GRID', (0,-3), (-1,-1), 0.5, colors.black), # Grid for the last row
        ('SPAN', (1,-1), (4,-1)),  # Span the description cell for the final total
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),  # Right align quantities and prices
        ('ALIGN', (0,1), (0,-1), 'CENTER'),  # Center align item numbers
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    # === Recap Section ===
    elements.append(Paragraph("<b>RÉCAPITULATION GÉNÉRALE</b>", style_bold))
    elements.append(Spacer(1, 6))
    previous_total = (
    Decompte.objects
    .filter(
        marche=decompte.marche,
        created_at__lt=decompte.created_at  # or id__lt=decompte.id
    )
    .aggregate(total=Sum('montant_ttc'))['total']
    or Decimal('0.00')
   )
    periodicite = decompte.periodicite or 1
    decompte_tva = tva/Decimal(periodicite)
    recap_data = [
        ["NATURES DES DÉPENSES", "DÉPENSES FAITES"],
        ["Prestation réalisées", f"{decompte.montant_ttc:.2f}"],
        ["TOTAUX TTC", f"{decompte.montant_ttc:.2f}"],
        ["À déduire le montant des acomptes délivrés sur l'exercice en cours", f"{previous_total:.2f}"],
        ["Montant de l'acompte à délivrer en DH TTC", f"{decompte.montant_ttc:.2f}"],
        ["Dont TVA (à 20%)", f"{decompte_tva:.2f}"]
    ]

    recap_table = Table(recap_data, colWidths=[13*cm, 6*cm])
    recap_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (1,1), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements.append(recap_table)
    elements.append(Spacer(1, 6))

    # === TVA Line ===
    #tva_line = Paragraph(f"Arrété le présent décompte {decompte.type} ,{decompte.numero} , par nous sous-ordonateur , a la somme de : {num2words(decompte.montant_ttc,lang='fr')} Dirhams Tous Taxes Comprises dont TVA {(decompte.montant_ttc * Decimal('0.2')):.2f} dhs", styles['Normal'])
    #elements.append(tva_line)
    montant = decompte.montant_ttc

# 1. Extract integer and decimal parts
    dirhams = int(montant)
    centimes = int((montant - Decimal(dirhams)) * 100)

        # 2. Convert each part to words
    dirhams_words = num2words(dirhams, lang='fr')
    centimes_words = num2words(centimes, lang='fr') if centimes > 0 else ""

        # 3. Build the phrase according to the rules
    if centimes > 0:
            montant_text = f"{dirhams_words} Dirhams et {centimes_words} Centimes"
    else:
            montant_text = f"{dirhams_words} Dirhams"

# 4. TVA numeric value
    tva_value = (montant * Decimal('0.2')).quantize(Decimal("0.01"))

        # 5. Full paragraph
    tva_line = Paragraph(
            f"Arrété le présent décompte {decompte.type}, {decompte.numero}, "
            f"par nous sous-ordonateur, à la somme de : {montant_text} "
            f"Toutes Taxes Comprises dont TVA {tva_value} dhs",
            styles['Normal']
        )

    elements.append(tva_line)
    # Build PDF
    doc.build(elements)
    return response
@login_required
def decompte_detail(request, pk):
    decompte = get_object_or_404(Decompte, pk=pk)
    return render(request, 'markets/decompte_detail.html', {
        'decompte': decompte,
        'marche': decompte.marche
    })


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from .models import Acompte, Marche

########"""

# Order SErvice Views

@login_required
def order_list(request):
    
    #decompte = Decompte.objects.prefetch_related('marche').all()
    order = OrdreService.objects.select_related('marche').all()
    #marches = Marche.objects.prefetch_related('decomptes').all()  
    marches = Marche.objects.select_related('order_services').all()
   
    paginator = Paginator(order, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/order_service_list.html', {
        'page_obj': page_obj,          # Paginated Decomptes
        'all_marches': marches,        # All Marches with related Decomptes
        'statut_choices': OrdreService.STATUT_CHOICES,
    })




@login_required
def order_create(request):
    if request.method == 'POST':
        form = OrdreServiceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ordre de service créé avec succès.')
            return redirect('order_list')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = OrdreServiceForm()

    return render(request, 'markets/order_form.html', {
        'form': form,
        'title': 'Nouveau ordre de service'
    })



@login_required
def order_detail(request, pk):
    order = get_object_or_404(OrdreService, pk=pk)
    return render(request, 'markets/order_detail.html', {
        'order': order,
        'marche': OrdreService.marche
    })

@login_required
def order_update(request, pk):
    order = get_object_or_404(OrdreService, pk=pk)
    if request.method == 'POST':
        form = OrdreServiceForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            messages.success(request, 'Order de sevice mis à jour avec succès.')
            return redirect('order_detail', pk=pk)
    else:
        form = OrdreServiceForm(instance=order)
    return render(request, 'markets/order_form.html', {'form': form, 'title': 'Modifier Order de Service'})


@login_required
def order_delete(request, pk):
    order = get_object_or_404(OrdreService, pk=pk)
    if request.method == 'POST':
       order.delete()
       messages.success(request, 'Marché supprimé avec succès.')
       return redirect('order_list')
    return render(request, 'markets/confirm_delete.html', {'object': order, 'type': 'Order de Service'})





@login_required
def generate_ordre_service_pdf(request, pk):
    ordre = get_object_or_404(OrdreService, pk=pk)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ordre_service_{ordre.numero}.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    normal = styles['Normal']
    bold = ParagraphStyle('Bold', parent=normal, fontName='Helvetica-Bold')
    centered = ParagraphStyle('Centered', parent=normal, alignment=1, fontSize=12, spaceAfter=10)
    paragraph_style = ParagraphStyle('Justify', parent=normal, alignment=4, leading=15, fontSize=11)

    elements = []

    # En-tête avec logos
    try:
        #logo_left = Image("static/img/logo1.png", width=3*cm, height=3*cm)
        logo_center = Image("static/img/LOGO.jpg", width=13*cm, height=3*cm)
        #logo_right = Image("static/img/logo3.png", width=3*cm, height=3*cm)
        header_table = Table([[logo_center]], colWidths=[5*cm, 5*cm, 5*cm])
        header_table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
        elements.append(header_table)
    except:
        pass

    elements.append(Spacer(1, 12))

    # Référence
    elements.append(Paragraph(f"Réf : {ordre.numero}", normal))
    elements.append(Spacer(1, 12))

    # Titre
    elements.append(Paragraph(f"<b>Ordre de Service N° {ordre.numero}</b>", centered))

    # === Paragraphe principal selon le type d'ordre ===
    if ordre.type == 'commencement':
        texte = f"""
        La Trésorerie Générale du Royaume, représentée par {ordre.marche.maitre_ouvrage or '________________'},
        invite la société <b>{ordre.marche.prestataire.nom if ordre.marche and ordre.marche.prestataire else '________________'}</b>,
        à exécuter à compter du {ordre.date_execution.strftime('%d/%m/%Y') if ordre.date_execution else '____/____/____'},
        les prestations relatives au marché <b>{ordre.marche.numero if ordre.marche else '____________'}</b>,
        concernant {ordre.objet}.
        """
    elif ordre.type == 'arret':
        texte = f"""
        La Trésorerie Générale du Royaume, représentée par {ordre.marche.maitre_ouvrage or '________________'},
        ordonne l’arrêt des prestations prévues dans le marché <b>{ordre.marche.numero if ordre.marche else '____________'}</b>,
        exécutées par la société <b>{ordre.marche.prestataire.nom if ordre.marche and ordre.marche.prestataire else '________________'}</b>,
        à compter du {ordre.date_execution.strftime('%d/%m/%Y') if ordre.date_execution else '____/____/____'}.
        """
    elif ordre.type == 'reprise':
        texte = f"""
        La Trésorerie Générale du Royaume, représentée par {ordre.marche.maitre_ouvrage or '________________'},
        ordonne la reprise des prestations relatives au marché <b>{ordre.marche.numero if ordre.marche else '____________'}</b>,
        confiées à la société <b>{ordre.marche.prestataire.nom if ordre.marche and ordre.marche.prestataire else '________________'}</b>,
        à compter du {ordre.date_execution.strftime('%d/%m/%Y') if ordre.date_execution else '____/____/____'}.
        """
    else:
        texte = f"""
        La Trésorerie Générale du Royaume, représentée par {ordre.marche.maitre_ouvrage or '________________'},
        informe la société <b>{ordre.marche.prestataire.nom if ordre.marche and ordre.marche.prestataire else '________________'}</b>
        concernant {ordre.objet}.
        """

    # Ajouter le paragraphe choisi
    elements.append(Paragraph(texte, paragraph_style))
    elements.append(Spacer(1, 20))

    # Date et signature (partie 1)
    elements.append(Paragraph(f"{ordre.date_emission.strftime('%d/%m/%Y') if ordre.date_emission else '____/____/____'}", normal))
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Signature et cachet", normal))
    elements.append(Spacer(1, 30))

    # Ligne de séparation
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<b>Notification</b>", centered))

    # Notification texte
    notif_text = f"""
    Je soussigné(e) {ordre.marche.prestataire.representant}, représentant la société <b>{ordre.marche.prestataire.nom if ordre.marche and ordre.marche.prestataire else '________________'}</b>,
    reconnais avoir reçu de la Trésorerie Générale du Royaume un exemplaire de l'ordre de service
    concernant le marché <b>{ordre.marche.numero if ordre.marche else '____________'}</b>.
    """
    elements.append(Paragraph(notif_text, paragraph_style))
    elements.append(Spacer(1, 20))

    # Date et signature (partie 2)
    elements.append(Paragraph(f"{ordre.date_emission.strftime('%d/%m/%Y') if ordre.date_emission else '____/____/____'}", normal))
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Signature et cachet du prestataire", normal))

    # Génération PDF
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
def pv_create(request):
    if request.method == 'POST':
        form = PVForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'PV créé avec succès.')
            return redirect('pv_list')
    else:
        form = PVForm()
    return render(request, 'markets/pv_form.html', {'form': form, 'title': 'Nouveau PV'})

@login_required
def pv_list(request):
    
    #decompte = Decompte.objects.prefetch_related('marche').all()
    pv = PV.objects.select_related('marche').all()
    #marches = Marche.objects.prefetch_related('decomptes').all()  
    marches = Marche.objects.select_related('pv').all()
   
    paginator = Paginator(pv, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/pv_list.html', {
        'page_obj': page_obj,          # Paginated Decomptes
        'all_marches': marches,        # All Marches with related Decomptes
        'statut_choices': OrdreService.STATUT_CHOICES,
    })


@login_required
def pv_update(request, pk):
    pv = get_object_or_404(PV, pk=pk)
    if request.method == 'POST':
        form = PVForm(request.POST, instance=pv)
        if form.is_valid():
            form.save()
            messages.success(request, 'PV  mis à jour avec succès.')
            return redirect('pv_detail', pk=pk)
    else:
        form = PVForm(instance=pv)
    return render(request, 'markets/pv_form.html', {'form': form, 'title': 'Modifier PV'})


@login_required
def pv_detail(request, pk):
    pv = get_object_or_404(PV, pk=pk)
    return render(request, 'markets/pv_detail.html', {
        'pv': pv,
        'marche': PV.marche
    })

@login_required
def pv_delete(request, pk):
    pv = get_object_or_404(PV, pk=pk)
    if request.method == 'POST':
       pv.delete()
       messages.success(request, 'PV supprimé avec succès.')
       return redirect('pv_list')
    return render(request, 'markets/confirm_delete.html', {'object': pv, 'type': 'pv'})



def date_en_lettres(date_obj):
    jour = num2words(date_obj.day, lang='fr')
    mois = date_obj.strftime('%B')  # mois en lettres (janvier, février…)
    annee = num2words(date_obj.year, lang='fr')
            # Capitaliser le mois (optionnel : français = minuscules normalement)
    mois = mois.lower()
    jour =jour 
    annee = annee
    return f"{jour} {mois} {annee}"

@login_required
def generate_pv_pdf(request, pk):
    pv = get_object_or_404(PV, pk=pk)

    # Préparer la réponse HTTP
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="pv_{pv.numero}.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    elements = []
    styles = getSampleStyleSheet()

    # === Styles personnalisés ===
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=14,
        spaceAfter=20,
        leading=18,
    )
    justify_style = ParagraphStyle(
        "Justify",
        parent=styles["Normal"],
        alignment=TA_JUSTIFY,
        fontSize=11,
        leading=16,
    )
    
    
    # === Logos en haut ===
    #logo1 = Image("static/img/logo1.png", width=100, height=100)
    logo2 = Image("static/img/LOGO.jpg", width=400, height=100)
    #logo3 = Image("static/img/logo3.png", width=100, height=100)

    logos_table = Table([[logo2]], colWidths=[150, 150, 150])
    logos_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
            ]
        )
    )
    elements.append(logos_table)
    elements.append(Spacer(1, 20))
    signatairess = pv.signataires.all()


    # === Choix du contenu selon le type de PV ===

    date_lettres = date_en_lettres(pv.date_pv)
    sinataires = pv.signataires.all()
    if pv.type == "reception provisoire":
        title = f"PROCÈS VERBAL DE RÉCEPTION PROVISOIRE<br/>RELATIF AU MARCHÉ   {pv.marche.numero}"
        corps = f"""
        Le {date_lettres} ; la commission chargée de la réception provisoire du marché  {pv.marche.numero}, 
        relatif à '{pv.marche.objet}', composée de :<br/><br/>
        """
        for sinataire in sinataires:
            corps+= f"-{sinataire.nom} ({sinataire.fonction})<br/>"
        corps += f"""<br/><br/>
    
        a constaté que les prestations exécutées par la société {pv.marche.prestataire} sont terminées et conformes.<br/><br/>
        En conséquence, la réception provisoire est prononcée.<br/><br/>
        <b>Fait à Rabat, le {pv.date_pv.strftime('%d/%m/%Y')}</b>.
        """

    elif pv.type == "reception defintive":
        title = f"PROCÈS VERBAL DE RÉCEPTION DÉFINITIVE<br/>RELATIF AU MARCHÉ {pv.marche.type}  {pv.marche.numero}"
        corps = f"""
        Le {date_lettres} ; la commission chargée de la réception définitive du marché  {pv.marche.numero}, 
        relatif à '{pv.marche.objet}', composée de :<br/><br/>
        """
        for sinataire in sinataires:
            corps+= f"-{sinataire.nom} ({sinataire.fonction})<br/>"
        corps += f""" <br/><br/>
        
        a reconnu que toutes les prestations exécutées par la société {pv.marche.prestataire} sont conformes aux conditions du marché.<br/><br/>
        En conséquence, la réception définitive est prononcée.<br/><br/>
        <b>Fait à Rabat, le {pv.date_pv.strftime('%d/%m/%Y')}</b>.
        """
 
    elif pv.type == "reception defintive parcielle":
        title = f"PROCÈS VERBAL DE RÉCEPTION DÉFINITIVE PARTIELLE<br/>RELATIF AU MARCHÉ {pv.marche.type}  {pv.marche.numero}"
        corps = f"""
        Le {date_lettres} ; la commission chargée de la réception définitive partielle du marché  {pv.marche.numero}, 
        relatif à ' {pv.marche.objet}', composée de :<br/><br/>
        """
        for sinataire in sinataires:
            corps+= f"-{sinataire.nom} ({sinataire.fonction})<br/>"
        corps += f""" <br/><br/>
        
        
        a reconnu que les prestations exécutées par la société {pv.marche.prestataire}, 
        pour la période {pv.periode_debut or ''} au {pv.periode_fin or ''}, 
        sont conformes aux conditions du marché.<br/><br/>
        En conséquence, la réception définitive partielle est prononcée.<br/><br/>
        <b>Fait à Rabat, le {pv.date_pv.strftime('%d/%m/%Y')}</b>.
        """

    else:
        title = f"PROCÈS VERBAL<br/>Marché N° {pv.numero}"
        corps = f"""
        Ce procès-verbal a été établi le {pv.date_pv.strftime('%d %B %Y')} 
        concernant le marché N° {pv.numero}, relatif à {pv.objet}.
        """

    # === Ajouter titre et corps ===
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(corps, justify_style))
    elements.append(Spacer(1, 30))

    # === Signatures ===
    elements.append(Paragraph("<b>Signé :</b>", styles["Normal"]))
    elements.append(Spacer(1, 12))

    signataires = []
    signatairess = pv.signataires.all()
    
    if signatairess.exists():
        for signataire in signatairess:
            signataires.append([f"• {signataire.nom}", ""])
    
   
    if signataires:
        table = Table(signataires, colWidths=[250, 250])
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 11),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
                ]
            )
        )
        elements.append(table)

    # === Génération du PDF ===
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


@login_required
def document_list(request):
    
    #decompte = Decompte.objects.prefetch_related('marche').all()
    document = Document.objects.select_related('marche').all()
    #marches = Marche.objects.prefetch_related('decomptes').all()  
    marches = Marche.objects.select_related('document').all()
   
    paginator = Paginator(document, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/document_list.html', {
        'page_obj': page_obj,          # Paginated Decomptes
        'all_marches': marches,        # All Marches with related Decomptes
        'statut_choices': OrdreService.STATUT_CHOICES,
    })



@login_required
def document_create(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Document créé avec succès.')
            return redirect('document_list')
    else:
        form = DocumentForm()
    return render(request, 'markets/document_form.html', {'form': form, 'title': 'Nouveau Document'})

  
@login_required
def document_update(request, pk):
    document = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        form = DocumentForm(request.POST, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, 'Document  mis à jour avec succès.')
            return redirect('pv_detail', pk=pk)
    else:
        form = DocumentForm(instance=document)
    return render(request, 'markets/document_form.html', {'form': form, 'title': 'Modifier Document'})
@login_required
def document_detail(request, pk):
    document = get_object_or_404(Document, pk=pk)
    return render(request, 'markets/document_detail.html', {
        'document': document,
        'marche': Document.marche
    })

@login_required
def document_delete(request, pk):
    document = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
       document.delete()
       messages.success(request, 'Document supprimé avec succès.')
       return redirect('document_list')
    return render(request, 'markets/confirm_delete.html', {'object': document, 'type': 'document'})


@login_required
def acompte_list(request):
    try:
        # Get all acomptes with their related marche
        acomptes = Acompte.objects.select_related('marche').all()
        
        # Pagination
        paginator = Paginator(acomptes, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        hi = 'hi'

        return render(request, 'markets/acompte_list.html', {
            'page_obj': page_obj,
            'hi' :hi
        })
    except Exception as e:
        print(f"Error in acompte_list: {e}")  # Voir l'erreur dans la console
        # Pour debug, retourner une réponse simple
        from django.http import HttpResponse
        return HttpResponse(f"Error: {e}")


@login_required
def acompte_create(request):
    if request.method == 'POST':
        form = AcompteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Acompte créé avec succès.')
            return redirect('acompte_list')
    else:
        form = AcompteForm()
    return render(request, 'markets/acompte_form.html', {'form': form, 'title': 'Nouveau Acompte'})

@login_required
def acompte_detail(request, pk):
    acompte = get_object_or_404(Acompte, pk=pk)
    return render(request, 'markets/acompte_detail.html', {
        'acompte': acompte,
        'marche': Acompte.marche
    })


    
@login_required
def acompte_update(request, pk):
    acompte = get_object_or_404(Acompte, pk=pk)
    if request.method == 'POST':
        form = AcompteForm(request.POST, instance=acompte)
        if form.is_valid():
            form.save()
            messages.success(request, 'Acompte  mis à jour avec succès.')
            return redirect('acompte_detail', pk=pk)
    else:
        form = AcompteForm(instance=acompte)
    return render(request, 'markets/acompte_form.html', {'form': form, 'title': 'Modifier acompte'})


@login_required
def acompte_delete(request, pk):
    acompte = get_object_or_404(Acompte, pk=pk)
    if request.method == 'POST':
       acompte.delete()
       messages.success(request, 'Acompte supprimé avec succès.')
       return redirect('acompte_list')
    return render(request, 'markets/confirm_delete.html', {'object': acompte, 'type': 'document'})

############################################## Signataires ##################################################################################


@login_required
def signataire_create(request):
    if request.method == 'POST':
        form = SignataireForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Signataire créé avec succès.')
            return redirect('signataire_list')
    else:
        form = SignataireForm()
    return render(request, 'markets/signataire_form.html', {'form': form, 'title': 'Nouveau Signataire'})




@login_required
def signataire_list(request):
    try:
        # Get all acomptes with their related marche
        signataire = Signataire.objects.all()
        
        # Pagination
        paginator = Paginator(signataire, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        

        return render(request, 'markets/signataire_list.html', {
            'page_obj': page_obj,
            
        })
    except Exception as e:
        print(f"Error in signataire_list: {e}")  # Voir l'erreur dans la console
        # Pour debug, retourner une réponse simple
        from django.http import HttpResponse
        return HttpResponse(f"Error: {e}")


@login_required
def signataire_delete(request, pk):
    signataire = get_object_or_404(Signataire, pk=pk)
    if request.method == 'POST':
       signataire.delete()
       messages.success(request, 'Signataire supprimé avec succès.')
       return redirect('signataire_list')
    return render(request, 'markets/confirm_delete.html', {'object': signataire, 'type': 'document'})

@login_required
def signataire_update(request, pk):
    signataire = get_object_or_404(Signataire, pk=pk)
    if request.method == 'POST':
        form = SignataireForm(request.POST, instance=signataire)
        if form.is_valid():
            form.save()
            messages.success(request, 'Signataire  mis à jour avec succès.')
            return redirect('signataire_detail', pk=pk)
    else:
        form = SignataireForm(instance=signataire)
    return render(request, 'markets/signataire_form.html', {'form': form, 'title': 'Modifier signataire'})

@login_required
def signataire_detail(request, pk):
    signataire = get_object_or_404(Signataire, pk=pk)
    return render(request, 'markets/signataire_detail.html', {
        'signataire': signataire,
        
    })







from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Ligne, Marche
from .forms import LigneForm

@login_required
def ligne_list(request):
    
    #decompte = Decompte.objects.prefetch_related('marche').all()
    ligne = Ligne.objects.select_related('marche').all()
    #marches = Marche.objects.prefetch_related('decomptes').all()  
    lignes = Marche.objects.select_related('lignes').all()
   
    paginator = Paginator(ligne, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'markets/ligne_list.html', {
        'page_obj': page_obj,          # Paginated Decomptes
        'all_marches': lignes,        # All Marches with related Decomptes
        
    })


@login_required
def ligne_create(request):
    if request.method == 'POST':
        form = LigneForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ligne créé avec succès.')
            return redirect('ligne_list')
    else:
        form = LigneForm()
    return render(request, 'markets/pv_form.html', {'form': form, 'title': 'Nouveau Ligne'})



@login_required
def ligne_update(request, pk):
    ligne = get_object_or_404(Ligne, pk=pk)
    if request.method == 'POST':
        form = LigneForm(request.POST, instance=ligne)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ligne mis à jour avec succès.')
            return redirect('ligne_detail', pk=pk)
    else:
        form = LigneForm(instance=ligne)
    return render(request, 'markets/ligne_form.html', {'form': form, 'title': 'Modifier Ligne'})

@login_required
def ligne_delete(request, pk):
    ligne = get_object_or_404(Ligne, pk=pk)
    if request.method == 'POST':
       ligne.delete()
       messages.success(request, 'Ligne supprimé avec succès.')
       return redirect('ligne_list')
    return render(request, 'markets/confirm_delete.html', {'object': ligne, 'type': 'ligne'})

@login_required
def ligne_detail(request, pk):
    ligne = get_object_or_404(Ligne, pk=pk)
    return render(request, 'markets/ligne_detail.html', {
        'ligne': ligne,
        'marche': Ligne.marche
    })
